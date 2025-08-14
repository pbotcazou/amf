# app_streamlit_amf_v4.py
# -*- coding: utf-8 -*-

import json
import random
import re
from pathlib import Path
from typing import List, Optional

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

# ==================== Configuration ====================
APP_TITLE    = "Entraînement Certification AMF"
DEFAULT_XLSX = "AMF.xlsx"        # Mets ici le nom exact de ton fichier (ou importe via la barre latérale)
SHEET_NAME   = "V4"

# Tailles des sessions
QUIZ_SIZE     = 84               # Mode examen aléatoire
BATCH_SIZE    = 20               # Mode parcours 20 par 20
SPRINT_BATCH  = 2                # 2 questions par mini-batch
SPRINT_COUNT  = 7                # 7 mini-batches => 14 questions

# Fichiers de persistance (conservés entre exécutions)
SEEN_FILE       = Path(".amf_seen_ids.json")      # questions vues (après correction)
WRONG_FILE      = Path(".amf_wrong_ids.json")     # questions en erreur
PROGRESS_FILE   = Path(".amf_progress.json")      # état du parcours 20x20 (ordre + curseur)
SPRINT_FILE     = Path(".amf_sprint.json")        # état du sprint (ordre aléatoire + curseur)

# Compat Streamlit rerun
try:
    RERUN = st.rerun
except AttributeError:
    RERUN = st.experimental_rerun  # type: ignore[attr-defined]


# ==================== Helpers ====================
def s(val) -> str:
    return "" if val is None else str(val).strip()

def clean_question_text(txt: str) -> str:
    """
    Supprime un éventuel préfixe '123 - ' dans l'énoncé.
    Ex: '961 - Le PSI ...' -> 'Le PSI ...'
    """
    if not txt:
        return ""
    return re.sub(r"^\s*\d+\s*-\s*", "", txt).strip()

def cell_is_yellow(cell) -> bool:
    """
    Détecte si la cellule est surlignée en jaune (plusieurs formats Excel possibles).
    """
    f = cell.fill
    if not f or not f.fgColor:
        return False

    rgb_val = getattr(f.fgColor, "rgb", None)
    if rgb_val:
        rgb_str = str(rgb_val).upper()
        for pat in ["FFEB9C", "FFFF00", "FFFDEB", "FFF2CC", "FFFFFF00"]:
            if pat in rgb_str:
                return True

    # Palette indexée (legacy)
    if getattr(f.fgColor, "indexed", None) is not None:
        return f.fgColor.indexed in (5, 6, 13, 27, 44)

    return False

def load_json(file_path: Path, default):
    if file_path.exists():
        try:
            return json.loads(file_path.read_text(encoding="utf-8"))
        except Exception:
            return default
    return default

def save_json(file_path: Path, obj) -> None:
    try:
        file_path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass

def load_json_ids(file_path: Path) -> set:
    data = load_json(file_path, [])
    return set(data) if isinstance(data, list) else set()

def save_json_ids(file_path: Path, ids: set) -> None:
    save_json(file_path, sorted(list(ids)))

def load_seen_ids() -> set:
    return load_json_ids(SEEN_FILE)

def save_seen_ids(seen: set) -> None:
    save_json_ids(SEEN_FILE, seen)

def load_wrong_ids() -> set:
    return load_json_ids(WRONG_FILE)

def save_wrong_ids(wrong: set) -> None:
    save_json_ids(WRONG_FILE, wrong)

def load_progress() -> dict:
    # 20x20 : {"order":[ids...], "cursor": int}
    return load_json(PROGRESS_FILE, {"order": [], "cursor": 0})

def save_progress(order: List[int], cursor: int) -> None:
    save_json(PROGRESS_FILE, {"order": order, "cursor": cursor})

def load_sprint() -> dict:
    # Sprint : {"order":[ids...], "cursor": int}
    return load_json(SPRINT_FILE, {"order": [], "cursor": 0})

def save_sprint(order: List[int], cursor: int) -> None:
    save_json(SPRINT_FILE, {"order": order, "cursor": cursor})

def pick_quiz_ids(all_ids: List[int], seen: set, k: int) -> List[int]:
    """
    Privilégie les ids jamais vus, puis complète avec du random si pas assez.
    """
    unseen = [i for i in all_ids if i not in seen]
    random.shuffle(unseen)
    chosen = unseen[:k]
    if len(chosen) < k:
        remaining = [i for i in all_ids if i not in chosen]
        random.shuffle(remaining)
        chosen += remaining[:(k - len(chosen))]
    return chosen


# ==================== Chargement base ====================
@st.cache_data(show_spinner=False)
def load_questions_from_excel(xlsx_path: str, sheet_name: str = SHEET_NAME) -> pd.DataFrame:
    """
    Retourne un DataFrame:
      id, question, A, B, C, correct_idx, correct_text
    La bonne réponse est repérée via le surlignage jaune.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Onglet introuvable: {sheet_name}")
    ws = wb[sheet_name]

    # Trouver la ligne après "n°identifiant"
    start_row = 1
    for i in range(1, ws.max_row + 1):
        v = ws.cell(row=i, column=1).value
        if v and s(v).lower().startswith("n°identifiant"):
            start_row = i + 1
            break

    records = []
    for r in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        rid = r[0].value
        if rid is None or (isinstance(rid, str) and rid.strip() == ""):
            continue
        try:
            rid_int = int(s(rid))
        except Exception:
            continue

        q_text_raw = s(r[1].value)
        q_text = clean_question_text(q_text_raw)  # Nettoyage de la double numérotation
        A = s(r[2].value)
        B = s(r[3].value)
        C = s(r[4].value)

        # Détection de la bonne réponse via la couleur
        correct_idx = None
        for idx, c in enumerate([r[2], r[3], r[4]]):
            try:
                if cell_is_yellow(c):
                    correct_idx = idx
                    break
            except Exception:
                pass

        correct_text = [A, B, C][correct_idx] if correct_idx is not None else ""
        if q_text:
            records.append({
                "id": rid_int, "question": q_text,
                "A": A, "B": B, "C": C,
                "correct_idx": correct_idx,
                "correct_text": correct_text
            })

    df = pd.DataFrame.from_records(records)
    if not df.empty:
        df = df[df["question"] != ""]
    return df


# ==================== Thème sombre (forcé) ====================
def style_dark():
    css_vars = dict(
        bg="#0b1020", text="#e5e7eb", card="#111827", border="rgba(255,255,255,.15)",
        correct_bg="rgba(16,185,129,0.18)", correct_border="#34d399",
        wrong_bg="rgba(239,68,68,0.20)", wrong_border="#f87171",
        muted="#9ca3af"
    )
    st.markdown(
        f"""
        <style>
        .main .block-container {{max-width: 980px;}}
        html, body, [data-testid="stAppViewContainer"] {{
            background: {css_vars['bg']} !important;
            color: {css_vars['text']} !important;
        }}
        .question-card {{
            border: 1px solid {css_vars['border']};
            border-radius: 14px;
            padding: 16px 18px;
            margin-bottom: 14px;
            background: {css_vars['card']};
            box-shadow: 0 1px 3px rgba(0,0,0,0.06);
        }}
        .question-title {{ font-weight: 700; margin-bottom: 10px; }}
        .qid-badge {{
            font-size: 12px; color: {css_vars['muted']}; margin-top: -6px; margin-bottom: 6px;
        }}
        .correct {{
            background: {css_vars['correct_bg']};
            border-left: 6px solid {css_vars['correct_border']};
            padding: 12px 14px; border-radius: 10px; color: inherit;
        }}
        .wrong {{
            background: {css_vars['wrong_bg']};
            border-left: 6px solid {css_vars['wrong_border']};
            padding: 12px 14px; border-radius: 10px; color: inherit;
        }}
        .muted {{ color: {css_vars['muted']}; }}
        .pill {{
            display: inline-block; padding: 2px 8px; border-radius: 999px;
            font-size: 11px; font-weight: 700; border: 1px solid {css_vars['border']};
            margin-right: 6px;
        }}
        .ans-line {{ margin: 2px 0 2px 0; }}
        .ans-label {{ font-weight: 700; width: 28px; display: inline-block; }}
        </style>
        """,
        unsafe_allow_html=True
    )


# ==================== UI ====================
def render_header():
    st.title(APP_TITLE)
    st.caption("• Examen 84 aléatoires  • Parcours 20×20 séquentiel  • Sprint 7×2 aléatoires  • Erreurs persistantes")

def sidebar_controls(df: pd.DataFrame):
    st.sidebar.header("Paramètres")

    # Choix du mode (3 modes)
    modes = ["Examen 84 aléatoires", "Parcours 20 par 20", "Sprint 7×2 aléatoires"]
    mode = st.sidebar.radio("Mode d'entraînement", modes, index=0)
    # Reset propre quand le mode change
    prev_mode = st.session_state.get("mode_prev")
    if prev_mode is None:
        st.session_state["mode_prev"] = mode
    elif prev_mode != mode:
        for k in ["quiz_started", "submitted", "review_wrong_only", "quiz_ids", "answers",
                  "cursor", "order_len", "mark_review"]:
            st.session_state.pop(k, None)
        st.session_state["mode_prev"] = mode
    st.session_state["mode"] = mode

    # Uploader Excel
    uploaded = st.sidebar.file_uploader("Importer un fichier Excel AMF", type=["xlsx"])
    if uploaded is not None:
        tmp_path = Path("uploaded_amf.xlsx")
        with open(tmp_path, "wb") as f:
            f.write(uploaded.getbuffer())
        st.session_state["xlsx_path"] = str(tmp_path)
        st.sidebar.success("Fichier importé. Recharge en cours…")
        RERUN()

    st.sidebar.write("")
    if st.sidebar.button("🔁 Nouveau test", use_container_width=True):
        for k in ["quiz_started", "submitted", "review_wrong_only", "quiz_ids", "answers", "mark_review"]:
            st.session_state.pop(k, None)
        # On ne touche pas aux fichiers de progression ici
        RERUN()

    wrong_ids = load_wrong_ids()
    if st.sidebar.button(f"📌 Réviser mes erreurs ({len(wrong_ids)})", use_container_width=True):
        for k in ["quiz_started", "submitted"]:
            st.session_state.pop(k, None)
        st.session_state["mode"] = "Examen 84 aléatoires"
        st.session_state["mode_prev"] = "Examen 84 aléatoires"
        st.session_state["review_wrong_only"] = True
        RERUN()

    st.sidebar.write("---")
    if st.sidebar.button("🧹 Réinitialiser l'historique (vu)", use_container_width=True):
        if SEEN_FILE.exists(): SEEN_FILE.unlink(missing_ok=True)
        st.sidebar.success("Historique 'vu' effacé.")
        RERUN()

    if st.sidebar.button("🧹 Réinitialiser mes erreurs", use_container_width=True):
        if WRONG_FILE.exists(): WRONG_FILE.unlink(missing_ok=True)
        st.sidebar.success("Liste d'erreurs effacée.")
        RERUN()

    # 🔥 On enlève le bouton de réinitialisation du parcours 20×20 comme demandé

    st.sidebar.write("---")
    st.sidebar.caption(f"{len(df)} questions dans la base.")


# ==================== Démarrage des modes ====================
def start_quiz_examen(df: pd.DataFrame):
    all_ids = df["id"].tolist()
    if st.session_state.get("review_wrong_only"):
        wrong_ids = list(load_wrong_ids())
        chosen_ids = wrong_ids if len(wrong_ids) <= QUIZ_SIZE else random.sample(wrong_ids, QUIZ_SIZE)
    else:
        seen = load_seen_ids()
        chosen_ids = pick_quiz_ids(all_ids, seen, QUIZ_SIZE)

    st.session_state["quiz_ids"] = chosen_ids
    st.session_state["answers"] = {}            # id -> "A"/"B"/"C" ou None
    st.session_state["mark_review"] = {}        # id -> True/False
    st.session_state["submitted"] = False
    st.session_state["quiz_started"] = True


def start_quiz_parcours(df: pd.DataFrame):
    # Charger ou initialiser l'ordre + curseur
    prog = load_progress()
    order = prog.get("order", [])
    cursor = int(prog.get("cursor", 0))

    if not order:  # première fois -> ordre déterministe (tri par id)
        order = sorted(df["id"].tolist())
        cursor = 0
        save_progress(order, cursor)

    # Sélection du batch courant
    end = min(cursor + BATCH_SIZE, len(order))
    chosen_ids = order[cursor:end]

    st.session_state["quiz_ids"] = chosen_ids
    st.session_state["answers"] = {}
    st.session_state["mark_review"] = {}
    st.session_state["submitted"] = False
    st.session_state["quiz_started"] = True
    st.session_state["cursor"] = cursor         # mémorise pour ce batch
    st.session_state["order_len"] = len(order)


def start_quiz_sprint(df: pd.DataFrame):
    # Charger ou initialiser l'ordre aléatoire + curseur (pour 14 questions max)
    sp = load_sprint()
    order = sp.get("order", [])
    cursor = int(sp.get("cursor", 0))

    if not order:
        all_ids = df["id"].tolist()
        random.shuffle(all_ids)
        order = all_ids[:SPRINT_BATCH * SPRINT_COUNT]  # 14 max
        cursor = 0
        save_sprint(order, cursor)

    end = min(cursor + SPRINT_BATCH, len(order))
    chosen_ids = order[cursor:end]

    st.session_state["quiz_ids"] = chosen_ids
    st.session_state["answers"] = {}
    st.session_state["mark_review"] = {}
    st.session_state["submitted"] = False
    st.session_state["quiz_started"] = True
    st.session_state["sprint_cursor"] = cursor
    st.session_state["sprint_total"] = len(order)


# ==================== Rendu ====================
def render_progress_bar():
    ids = st.session_state.get("quiz_ids", [])
    answers = st.session_state.get("answers", {})
    answered = sum(1 for qid in ids if answers.get(qid) in ["A", "B", "C"])
    st.progress(answered / max(len(ids), 1))
    st.caption(f"Répondu : {answered} / {len(ids)}")


def render_quiz(df: pd.DataFrame):
    ids = st.session_state["quiz_ids"]
    answers = st.session_state["answers"]
    marks   = st.session_state["mark_review"]
    rows = df.set_index("id").loc[ids].reset_index()

    render_progress_bar()

    for i, row in rows.iterrows():
        qid = int(row["id"])
        options = [f"A) {row['A']}", f"B) {row['B']}", f"C) {row['C']}"]

        st.markdown("<div class='question-card'>", unsafe_allow_html=True)
        st.markdown(f"<div class='question-title'>Q{i+1}/{len(ids)}. {row['question']}</div>", unsafe_allow_html=True)
        st.markdown(f"<div class='qid-badge'>ID: {qid}</div>", unsafe_allow_html=True)

        # Radio sans présélection : on ajoute un placeholder en première option
        radio_options = ["— Choisir —"] + options
        current = answers.get(qid)
        if current is None:
            index = 0
        else:
            index = {"A":1, "B":2, "C":3}[current]

        selection = st.radio(
            label="",
            options=radio_options,
            index=index,
            key=f"q_{qid}"
        )

        chosen_letter: Optional[str] = None
        if selection.startswith("A)"):
            chosen_letter = "A"
        elif selection.startswith("B)"):
            chosen_letter = "B"
        elif selection.startswith("C)"):
            chosen_letter = "C"
        else:
            chosen_letter = None

        st.session_state["answers"][qid] = chosen_letter

        # Marquer à revoir (ajoute à la liste d’erreurs même si bon)
        mark = st.checkbox("📌 Marquer à revoir", key=f"mark_{qid}", value=marks.get(qid, False))
        st.session_state["mark_review"][qid] = mark

        st.markdown("</div>", unsafe_allow_html=True)

    st.write("")
    if st.button("🟡 Voir les réponses", type="primary", use_container_width=True):
        st.session_state["submitted"] = True


def grade_and_show_results(df: pd.DataFrame):
    ids = st.session_state.get("quiz_ids", [])
    answers = st.session_state.get("answers", {})
    marks   = st.session_state.get("mark_review", {})
    if not ids:
        st.warning("Aucune question chargée.")
        return

    rows = df.set_index("id").loc[ids].reset_index()

    score = 0
    details = []
    wrong_ids_set = load_wrong_ids()

    for _, row in rows.iterrows():
        qid = int(row["id"])
        correct_idx = row["correct_idx"]
        correct_letter = ["A", "B", "C"][correct_idx] if pd.notna(correct_idx) and correct_idx in [0, 1, 2] else None
        user_letter = answers.get(qid)  # None si non répondu

        correct = (user_letter is not None and user_letter == correct_letter)
        if correct:
            score += 1
            if qid in wrong_ids_set:
                wrong_ids_set.remove(qid)
        else:
            wrong_ids_set.add(qid)

        # Forcer en "à revoir" si coché
        if marks.get(qid):
            wrong_ids_set.add(qid)

        details.append((qid, row["question"], user_letter, correct_letter, row["A"], row["B"], row["C"]))

    # Sauvegarde erreurs & “vues” (uniquement à la correction)
    save_wrong_ids(wrong_ids_set)

    seen = load_seen_ids()
    for (qid, _, _, correct_letter, *_rest) in details:
        if correct_letter is not None:
            seen.add(qid)
    save_seen_ids(seen)

    # En-tête score
    st.subheader("Résultats")
    st.markdown(f"### Score : **{score} / {len(ids)}**")

    # Afficher d'abord les erreurs, puis les bonnes
    details_sorted = sorted(details, key=lambda x: (x[2] == x[3], ))  # Faux/None en premier

    for (qid, qtext, user_letter, correct_letter, A, B, C) in details_sorted:
        if correct_letter is None:
            st.markdown(
                f"<div class='wrong'><b>ID {qid}.</b> {qtext}<br>"
                f"<span class='muted'>Impossible de détecter la bonne réponse (vérifie le surlignage jaune dans l’Excel).</span></div>",
                unsafe_allow_html=True
            )
            continue

        is_ok = (user_letter == correct_letter)
        box_class = "correct" if is_ok else "wrong"
        st.markdown(f"<div class='{box_class}'><b>ID {qid}.</b> {qtext}</div>", unsafe_allow_html=True)

        mapping = {"A": A, "B": B, "C": C}
        # Lignes de réponses : ✅ Bonne / 🔘 Ta réponse si fausse
        lines = []
        for letter in ["A", "B", "C"]:
            icon = "✅" if letter == correct_letter else ("🔘" if user_letter == letter and user_letter != correct_letter else "")
            text = mapping[letter] if mapping[letter] else "—"
            lines.append(f"<div class='ans-line'><span class='ans-label'>{letter})</span> {icon} {text}</div>")
        st.markdown("<div style='margin:6px 0 16px 12px;'>" + "".join(lines) + "</div>", unsafe_allow_html=True)

    # Contrôles de navigation pour Parcours & Sprint
    mode = st.session_state.get("mode")

    if mode == "Parcours 20 par 20":
        prog = load_progress()
        order = prog.get("order", [])
        cursor = int(prog.get("cursor", 0))
        st.info(f"Progression : {min(cursor + len(ids), len(order))} / {len(order)} si tu continues maintenant.")
        cols = st.columns(2)
        with cols[0]:
            if st.button("⬅️ Batch précédent", use_container_width=True) and cursor > 0:
                new_cursor = max(cursor - BATCH_SIZE, 0)
                save_progress(order, new_cursor)
                for k in ["quiz_started", "submitted"]:
                    st.session_state.pop(k, None)
                RERUN()
        with cols[1]:
            if st.button("➡️ Continuer (batch suivant)", use_container_width=True) and cursor < len(order):
                new_cursor = min(cursor + BATCH_SIZE, len(order))
                save_progress(order, new_cursor)
                for k in ["quiz_started", "submitted"]:
                    st.session_state.pop(k, None)
                RERUN()

    elif mode == "Sprint 7×2 aléatoires":
        sp = load_sprint()
        order = sp.get("order", [])
        cursor = int(sp.get("cursor", 0))
        total = len(order)
        st.info(f"Sprint : {min(cursor + len(ids), total)} / {total} si tu continues maintenant.")
        cols = st.columns(2)
        with cols[0]:
            if st.button("⬅️ Mini-batch précédent", use_container_width=True) and cursor > 0:
                new_cursor = max(cursor - SPRINT_BATCH, 0)
                save_sprint(order, new_cursor)
                for k in ["quiz_started", "submitted"]:
                    st.session_state.pop(k, None)
                RERUN()
        with cols[1]:
            if st.button("➡️ Continuer (mini-batch suivant)", use_container_width=True) and cursor < total:
                new_cursor = min(cursor + SPRINT_BATCH, total)
                save_sprint(order, new_cursor)
                for k in ["quiz_started", "submitted"]:
                    st.session_state.pop(k, None)
                RERUN()

    st.write("")
    if st.button("🔁 Relancer", use_container_width=True):
        for k in ["quiz_started", "submitted", "review_wrong_only"]:
            st.session_state.pop(k, None)
        RERUN()


# ==================== Main ====================
def main():
    style_dark()
    st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)  # petit spacer
    render_header()

    # Path Excel
    if "xlsx_path" not in st.session_state:
        st.session_state["xlsx_path"] = DEFAULT_XLSX

    xlsx_path = st.session_state["xlsx_path"]
    if not Path(xlsx_path).exists():
        st.info(f"Place le fichier **{DEFAULT_XLSX}** à côté du script ou importe-le via la barre latérale.")
        sidebar_controls(pd.DataFrame({"question": []}))
        return

    # Charger base
    try:
        df = load_questions_from_excel(xlsx_path, sheet_name=SHEET_NAME)
    except Exception as e:
        st.error(f"Impossible de lire le fichier Excel : {e}")
        return

    sidebar_controls(df)

    mode = st.session_state.get("mode", "Examen 84 aléatoires")

    # Démarrage du test suivant le mode
    if not st.session_state.get("quiz_started", False):
        if mode == "Examen 84 aléatoires":
            if len(df) < QUIZ_SIZE and not st.session_state.get("review_wrong_only", False):
                st.warning(f"La base ne contient que {len(df)} questions. Le test aura {len(df)} questions.")
            start_quiz_examen(df)

        elif mode == "Parcours 20 par 20":
            if len(df) < 1:
                st.warning("Aucune question disponible.")
                return
            start_quiz_parcours(df)

        else:  # Sprint 7×2 aléatoires
            if len(df) < 1:
                st.warning("Aucune question disponible.")
                return
            start_quiz_sprint(df)

    # Affichage
    if not st.session_state.get("submitted", False):
        render_quiz(df)
    else:
        grade_and_show_results(df)


if __name__ == "__main__":
    main()
